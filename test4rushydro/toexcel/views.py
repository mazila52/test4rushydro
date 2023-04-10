from django.http.response import HttpResponse
import pandas as pd

from pandas.core.frame import DataFrame
from django.core.files.storage import FileSystemStorage
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# Функция прнимает файл пользователя, запускает обработку и возвращает
def toexcel(request):
    if request.method == 'POST' and request.FILES['userfile'].content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        file = request.FILES['userfile']
        df = clear_xlsx(file)
        wb = send_xlsx(do_xlsx(), df)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment;filename="result.xlsx"'
        wb.save(response)
        return response
    return render(request, 'upload.html')


# Функция чистит файл пользователя и правит заголовки(для удобства)
def clear_xlsx(file) -> DataFrame:
    source_df = pd.read_excel(file)
    df = source_df[['Филиал', 'Сотрудник','Налоговая база','Налог']].copy()
    df = df.dropna(subset=['Филиал', 'Сотрудник'])
    df.columns = ['Филиал','Сотрудник','Налоговая база','Исчислено всего']
    df['Исчислено всего по формуле'] = [x * 0.13 if x < 5000000 else x * 0.15 for x in df['Налоговая база']]
    df['Отклонения'] = df['Исчислено всего'] - df['Исчислено всего по формуле']
    df = df.sort_values(by='Отклонения', key=lambda x: x.abs())
    return df


# Функция для создания xlsx файла по тербуемому шаблону
def do_xlsx() -> Workbook:
    # Создаем стиль ячеек заголовков
    header_style = NamedStyle(name='header')
    header_style.font = Font(name='Arial', size=10, bold=True)
    header_style.fill = PatternFill(fill_type='solid', fgColor='cbe4e5')
    header_style.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    header_style.border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000'),
        outline=Side(border_style='thin', color='FF000000'),
        vertical=Side(border_style='thin', color='FF000000'),
        horizontal=Side(border_style='thin', color='FF000000')
    )
    
    # Создаем рабочую область
    wb = Workbook()
    wb.add_named_style(header_style)
    ws = wb.active
    
    # Объединяем ячейки, настраиваем размеры полей
    ws.row_dimensions[1].height = 13
    ws.row_dimensions[2].height = 26
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 13.5
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 12.5
    ws['A1'] = 'Филиал'
    ws['A1'].style = 'header'
    ws['A2'].style = 'header'
    ws['B1'] = 'Сотрудник'
    ws['B1'].style = 'header'
    ws['C1'] = 'Налоговая база'
    ws['C1'].style = 'header'
    ws['D1'] = 'Налог'
    ws['D1'].style = 'header'
    ws['D2'] = 'Исчислено всего'
    ws['D2'].style = 'header'
    ws['E2'] = 'Исчислено всего по формуле'
    ws['E2'].style = 'header'
    ws['F1'] = 'Отклонения'
    ws['F1'].style = 'header'
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    ws.merge_cells('D1:E1')
    ws.merge_cells('F1:F2')
    return wb


# Функция вставляет обработанный DataFrame в созданный нами шаблон
def send_xlsx(wb: Workbook, df: DataFrame):
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)

    wb.save('static/result.xlsx')
    return wb
